from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


class Export:
    def __init__(self):
        Path('result data').mkdir(parents=True, exist_ok=True)
        # TODO: время конца парсинга в файл добавить
        self.start_time = self.get_start_time()
        self.f_name = f'result-{self.start_time}.xlsx'
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Sheet 1"
        self.sheet = self.wb['Sheet 1']
        self.wb.save(filename=Path('result data', self.f_name))
        self.create_titles()

    column_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'))

    @staticmethod
    def get_start_time():
        f_name = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        return f_name

    def create_titles(self):
        self.sheet.column_dimensions['A'].width = 7
        self.purchases_count = 1

        column_border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))

        self.set_styles('2', column_border)

        self.sheet[f'A1'].value = f'Начало парсинга: {self.start_time}'
        self.sheet.column_dimensions['A'].width = 10
        self.sheet[f'A2'].value = '№ номер'
        self.sheet.column_dimensions['B'].width = 15
        self.sheet[f'B2'].value = '№ аукциона'
        self.sheet.column_dimensions['C'].width = 15
        self.sheet[f'C2'].value = 'Статус контракта'
        self.sheet.column_dimensions['D'].width = 15
        self.sheet[f'D2'].value = 'Начало подачи заявок'
        self.sheet.column_dimensions['E'].width = 15
        self.sheet[f'E2'].value = 'Конец подачи заявок'
        self.sheet.column_dimensions['F'].width = 15
        self.sheet[f'F2'].value = 'Дата и время проведения'
        self.sheet.column_dimensions['G'].width = 15
        self.sheet[f'G2'].value = 'НМЦ'
        self.sheet.column_dimensions['H'].width = 15
        self.sheet[f'H2'].value = 'Регион'
        self.sheet.column_dimensions['I'].width = 80
        self.sheet[f'I2'].value = 'Заказчик'
        self.sheet.column_dimensions['J'].width = 10
        self.sheet[f'J2'].value = 'КТРУ'
        self.sheet.column_dimensions['J'].width = 20
        self.sheet[f'K2'].value = 'Наименование товара'
        self.sheet[f'L2'].value = 'Количество товара'
        self.sheet[f'M2'].value = 'Список участников'
        self.sheet[f'N2'].value = 'Сумма в заявке'
        self.sheet.column_dimensions['O'].width = 40
        self.sheet[f'O2'].value = 'Поставщик'
        self.sheet.column_dimensions['P'].width = 15
        self.sheet[f'P2'].value = 'Сумма по аукциону'
        self.sheet.column_dimensions['Q'].width = 12
        self.sheet[f'Q2'].value = 'Сумма контракта'
        self.sheet[f'R2'].value = 'Производитель товара'
        self.sheet[f'S2'].value = 'Страна происхождения'
        self.sheet.column_dimensions['T'].width = 12
        self.sheet[f'T2'].value = 'Номер РУ'
        self.sheet.column_dimensions['U'].width = 90
        self.sheet[f'U2'].value = 'Ссылка на РУ'
        self.sheet.column_dimensions['V'].width = 10
        self.sheet[f'V2'].value = 'Срок действия РУ'
        self.sheet.column_dimensions['W'].width = 13
        self.sheet[f'W2'].value = '№ Реестровой записи'
        self.sheet.column_dimensions['X'].width = 60
        self.sheet[f'X2'].value = 'Ссылка на выписку из реестра МинПромТорга'
        self.sheet[f'Y2'].value = 'Файлы'
        self.sheet[f'Z2'].value = 'Упоминание'
        self.sheet.column_dimensions['AA'].width = 15
        self.sheet[f'AA2'].value = 'Почта'
        self.sheet[f'AB2'].value = 'Телефон'
        self.sheet.column_dimensions['AB'].width = 15

        self.wb.save(filename=Path('result data', self.f_name))

    def dump_ru_data(self, row):
        if self.purchase_page.ru_numbers:
            for number in self.purchase_page.ru_numbers:
                if number:
                    self.sheet[f'T{row}'].value = number
                    if self.purchase_page.ru_data.get(number):
                        self.sheet[f'U{row}'].value = self.purchase_page.ru_data[number]['download_link']
                        self.sheet[f'V{row}'].value = self.purchase_page.ru_data[number]['the_term_of_the_certificate']
                        self.set_styles(row, self.column_border)

                row += 1

    def dump_registry_entry_data(self, row):
        if self.purchase_page.registry_entry_numbers_numbers:
            for registry_number in self.purchase_page.registry_entry_numbers_numbers:
                if registry_number:
                    self.sheet[f'W{row}'].value = registry_number
                    self.sheet[f'X{row}'].value = self.purchase_page.registry_entry_data[registry_number]['link']
                    self.set_styles(row, self.column_border)

                row += 1

    def dump_ktru(self, row):
        if self.purchase_page.ktru_blocks:
            for block in self.purchase_page.ktru_blocks:
                if block['ktru_position_code']:
                    self.sheet[f'J{row}'].value = block['ktru_position_code']
                self.set_styles(row, self.column_border)
                self.sheet[f'K{row}'].value = block['ktru_name_of_product_or_service']
                self.sheet[f'L{row}'].value = block['ktru_count']

                row += 1

    def dump_purchase_supplier_results(self, row):
        if self.purchase_page.purchase_supplier_results:
            for block in self.purchase_page.purchase_supplier_results:
                self.set_styles(row, self.column_border)
                self.sheet[f'O{row}'].value = block['supplier']
                self.sheet[f'Q{row}'].value = block['contract_price']
                if block['customer']:
                    self.sheet[f'I{row}'].value = block['customer']

                row += 1

    def set_styles(self, row, styles):
        if self.purchases_count % 2 == 0:
            color = 'C3FAD2'
        else:
            color = 'F5EEDA'

        for rows in self.sheet.iter_rows(min_row=int(row), max_row=self.sheet.max_row, min_col=self.sheet.min_column,
                                         max_col=self.sheet.max_column):
            for cell in rows:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = styles

    def dump_data(self, purchase_page, purchases_count):
        row = self.sheet.max_row + 1

        self.purchase_page = purchase_page
        self.purchases_count = purchases_count

        self.sheet[f'A{row}'].value = purchases_count
        self.sheet[f'B{row}'].value = purchase_page.purchase_number
        self.sheet[f'B{row}'].hyperlink = purchase_page.link
        self.sheet[f'C{row}'].value = purchase_page.status
        self.sheet[f'D{row}'].value = purchase_page.date_and_time_of_the_application_beginning
        self.sheet[f'E{row}'].value = purchase_page.date_and_time_of_the_application_deadline
        self.sheet[f'F{row}'].value = purchase_page.date_of_the_procedure_for_submitting_proposals
        self.sheet[f'G{row}'].value = purchase_page.starting_price
        self.sheet[f'H{row}'].value = purchase_page.region
        self.sheet[f'I{row}'].value = purchase_page.customer
        self.sheet[f'P{row}'].value = purchase_page.ktru_sum_cost
        self.sheet[f'AA{row}'].value = purchase_page.email
        self.sheet[f'AB{row}'].value = purchase_page.phone_number
        # TODO: что с поставщиком? Чем отличается от списка участников? Решил, что списка участников не будет прост)0)00)

        self.dump_ktru(row)
        self.dump_purchase_supplier_results(row)

        self.dump_ru_data(row)
        self.dump_registry_entry_data(row)

        self.set_styles(row, self.column_border)
        self.wb.save(filename=Path('result data', self.f_name))
